import openpyxl
excel_file = openpyxl.load_workbook('calendar.xlsx')

olymps = excel_file['Уровни']

# cur_sheet_active = excel_file.active

s = input().lower()
line = 0
while line == 0:
    for x in range(3, 22):
        if s in str(olymps.cell(row=x, column=2).value).lower():
            line = x
    if line == 0:
        print('Уточните название олимпиады и попробуйте ещё раз')
        s = input()
data = []
for x in range(1, 6):
    data.append(olymps.cell(row=line, column=x).value)
print(f'УРОВЕНЬ: {data[3]}\n')


