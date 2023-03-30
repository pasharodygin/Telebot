import telebot
import requests
from telebot import types
from bs4 import BeautifulSoup as BS
import config
import pymorphy2
import openpyxl

bot = telebot.TeleBot(config.TOKEN)
excel_file = openpyxl.load_workbook('calendar.xlsx')
olymps = excel_file['Уровни']
link = ''


def print_olymp(message: types.Message):
    mess1 = str()
    for x in range(3, 22):
        mess1 += f'\U0001F538 {olymps.cell(row=x, column=2).value}\n'
    bot.send_message(message.chat.id, mess1, parse_mode='html')


def print_day(n):
    morph = pymorphy2.MorphAnalyzer()
    day = morph.parse('день')[0]
    print(n, day.make_agree_with_number(n).word)


def parser(url):
    r = requests.get(url)
    soup = BS(r.text, 'html.parser')
    olymps = soup.find_all('ul', class_="events")
    only_text = [c for c in olymps]
    ans = []
    for olymp in only_text:
        date = olymp.find('div', class_="ecwd-date").text
        time = olymp.find('div', class_="ecwd-time").text
        name = olymp.find('div', class_="event-details-title").text
        if '-' in time:
            time = time[:8] + ' ' + time[8:]
        else:
            time = ' ' + time
        ans.append(date + name + time)
    return ans


def wrong_request(message: types.Message):
    kb = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=2)
    mess = 'Наш бот не может распознать ваше сообщение, убедитесь, что вы его правильно написали или вернитесь в главное меню'
    btn1 = types.KeyboardButton(text='Вернуться в главное меню')
    kb.add(btn1)
    bot.send_message(message.chat.id, mess, parse_mode='html', reply_markup=kb)


list_of_olymps = parser(config.URL1)


@bot.message_handler(commands=['start'])
def start(message):
    kb = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=2)
    btn1 = types.KeyboardButton(text='Узнать все про олимпиаду')
    btn2 = types.KeyboardButton(text='Узнать уровень ОЛИМПИАДЫ')
    btn4 = types.KeyboardButton(text='Весь список олимпиад на ТЕКУЩИЙ МЕСЯЦ')
    btn5 = types.KeyboardButton(text='Получить ссылку на сайт олимпиады')
    kb.add(btn1, btn2, btn4, btn5)
    mess = f'Привет, <b>{message.from_user.first_name}!</b>'
    mess1 = f'Используй <b>меню кнопок</b> для дальнейшей работы'
    bot.send_message(message.chat.id, mess, parse_mode='html', reply_markup=kb)
    bot.send_message(message.chat.id, mess1, parse_mode='html')


@bot.message_handler(func=lambda x: x.text == 'Получить ссылку на сайт этой олимпиады')
def get_link(mess):
    global link
    kb = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=2)
    btn = types.KeyboardButton(text='Вернуться в главное меню')
    kb.add(btn)
    bot.send_message(mess.chat.id, link, reply_markup=kb)


@bot.message_handler(commands=['website'])
def website(message):
    markup = types.InlineKeyboardMarkup()
    markup.add(types.InlineKeyboardButton("Перейти на сайт ИТМО", url="https://itmo.ru"))
    bot.send_message(message.chat.id, "Перейдите на сайт", parse_mode='html', reply_markup=markup)


@bot.message_handler(func=lambda x: x.text == 'Весь список олимпиад на ТЕКУЩИЙ МЕСЯЦ')
def tec_month(message: types.Message):
    ans = []
    kb = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=2)
    btn1 = types.KeyboardButton(text='Вернуться в главное меню')
    kb.add(btn1)
    for s in list_of_olymps:
        ans.append(s)
    if len(ans) == 0:
        bot.send_message(message.chat.id, '\u274CУВЫ! В это месяце нет олимпиад по информатике', reply_markup=kb)
    else:
        bot.send_message(message.chat.id, '\n'.join(ans), reply_markup=kb)


@bot.message_handler(func=lambda x: x.text == 'Вернуться в главное меню')
def return_(message: types.Message):
    kb = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=2)
    btn1 = types.KeyboardButton(text='Узнать все про олимпиаду')
    btn2 = types.KeyboardButton(text='Узнать уровень ОЛИМПИАДЫ')
    btn3 = types.KeyboardButton(text='Получить ссылку на сайт олимпиады')
    btn4 = types.KeyboardButton(text='Весь список олимпиад на ТЕКУЩИЙ МЕСЯЦ')
    kb.add(btn1, btn2, btn3, btn4)
    bot.send_message(message.chat.id, "Теперь вы находитесь в главном меню, продолжайте работу", reply_markup=kb)


@bot.message_handler(func=lambda msg: msg.text == "Узнать уровень ОЛИМПИАДЫ")
def get_lvl(message: types.Message):
    mess1 = 'Чтобы узнать уровень олимпиады, введите: \n<i>узнать уровень <b>название олимпиады</b></i>\nНапример: узнать уровень Высшая проба'
    bot.send_message(message.chat.id, mess1, parse_mode='html')


@bot.message_handler(func=lambda x: x.text == 'Узнать все про олимпиаду')
def get_inf(message):
    mess1 = f'Чтобы узнать всё про олимпиаду, введите: \n<i>узнать все <b>название олимпиады</b></i>\nНапример: узнать все Высшая проба'
    bot.send_message(message.chat.id, mess1, parse_mode='html')


@bot.message_handler(func=lambda x: x.text == 'Получить ссылку на сайт олимпиады')
def get_lol(message):
    mess1 = f'Чтобы получить ссылку олимпиады, введите: \n<i>ссылка <b>название олимпиады</b></i>\nНапример: ссылка Высшая проба'
    bot.send_message(message.chat.id, mess1, parse_mode='html')


@bot.message_handler(func=lambda msg: msg.text == "Вывести все олимпиады")
def if_sp(message: types.Message):
    print_olymp(message)


@bot.message_handler(content_types=['text'])
def olympiadas(message):
    text = message.text
    global link
    kb = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=2)
    if len(text) > 15 and text[:15].lower() == 'узнать уровень ':
        text = str(text[15:]).lower()
        line = 0
        for x in range(3, 22):
            if text in str(olymps.cell(row=x, column=2).value).lower():
                line = x
                break
        if line == 0:
            wrong_request(message)
        else:
            kb.row('Вернуться в главное меню', 'Получить ссылку на сайт этой олимпиады')
            k = olymps.cell(row=line, column=4).value
            link = str(olymps.cell(row=line, column=5).value)
            bot.send_message(message.chat.id, f'УРОВЕНЬ: {k}\n', reply_markup=kb)
    elif len(text) > 11 and text[:11].lower() == 'узнать все ':
        text = str(text[11:]).lower()
        line = 0
        for x in range(3, 22):
            if text in str(olymps.cell(row=x, column=2).value).lower():
                line = x
                break
        if line == 0:
            wrong_request(message)
        else:
            btn1 = types.KeyboardButton(text='Вернуться в главное меню')
            kb.add(btn1)
            mess = str(olymps.cell(row=line, column=2).value) + ':\n'
            mess += 'Номер в списке РСОШ: ' + str(olymps.cell(row=line, column=1).value) + '\n'
            mess += 'Ссылка на сайт: ' + str(olymps.cell(row=line, column=5).value) + '\n'
            mess += 'Уровень: ' + str(olymps.cell(row=line, column=4).value) + '\n'
            bot.send_message(message.chat.id, mess, reply_markup=kb)
    elif len(text) > 7 and text[:7].lower() == 'ссылка ':
        text = str(text[7:]).lower()
        line = 0
        for x in range(3, 22):
            if text in str(olymps.cell(row=x, column=2).value).lower():
                line = x
                break
        if line == 0:
            wrong_request(message)
        else:
            btn1 = types.KeyboardButton(text='Вернуться в главное меню')
            kb.add(btn1)
            mess = str(olymps.cell(row=line, column=5).value)
            bot.send_message(message.chat.id, mess, reply_markup=kb)
    else:
        wrong_request(message)


bot.polling()
